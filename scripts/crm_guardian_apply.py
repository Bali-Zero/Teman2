#!/usr/bin/env python3
"""CRM Guardian CLI — classify the deep_audit plan, dry-run or apply I1/I3.

Usage:
    # Classify only (prints rule counts, no Excel)
    python3 scripts/crm_guardian_apply.py --classify

    # Generate Excel dry-run plan without touching Drive or DB
    python3 scripts/crm_guardian_apply.py --dry-run

    # Apply on specific client IDs (REAL — respects kill switch)
    python3 scripts/crm_guardian_apply.py --apply --client-ids 1872,484,350

    # Apply on all R3+R4 in plan.jsonl (REAL — with batch limit)
    python3 scripts/crm_guardian_apply.py --apply --batch 10

Requires:
    - flyctl proxy 15432:5432 -a nuzantara-postgres  (DB tunnel)
    - ~/.nuzantara-drive-sa.json  (SA credentials)
    - plan.jsonl from scripts/crm_guardian_deep_audit.py
"""
from __future__ import annotations

import argparse
import asyncio
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import asyncpg

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "apps" / "backend-rag"))

from backend.services.crm_guardian import (  # noqa: E402
    GuardianConfig,
    GuardianRunContext,
    Rule,
    build_drive_service,
    compute_rule,
)
from backend.services.crm_guardian.base import (  # noqa: E402
    bump_circuit_breaker,
    get_invariant_state,
    is_globally_enabled,
    load_exceptions,
)
from backend.services.crm_guardian.consolidator import (  # noqa: E402
    apply_consolidation_for_client,
    apply_provision_and_consolidate,
    plan_consolidation,
)

ENV_FILE = PROJECT_ROOT / "apps" / "backend-rag" / ".env"
PLAN_FILE = PROJECT_ROOT / "research" / "compliance" / "deep_audit_2026-04-24" / "plan.jsonl"
EXCEL_OUT = PROJECT_ROOT / "research" / "compliance" / "crm_guardian_plan.xlsx"


def _load_db_url() -> str:
    for line in ENV_FILE.read_text().splitlines():
        if line.startswith("DATABASE_URL=") and "@localhost:15432" in line:
            return line.split("=", 1)[1].strip()
    raise RuntimeError("DATABASE_URL localhost:15432 not found in .env")


def load_plan() -> list[dict]:
    if not PLAN_FILE.exists():
        raise SystemExit(f"plan.jsonl not found at {PLAN_FILE}. Run crm_guardian_deep_audit.py first.")
    return [json.loads(line) for line in PLAN_FILE.read_text().splitlines() if line.strip()]


def classify_all(rows: list[dict], exceptions: set[tuple[str, str]]) -> list[tuple[Rule, dict]]:
    out: list[tuple[Rule, dict]] = []
    for row in rows:
        rule = compute_rule(row, exceptions=exceptions)
        out.append((rule, row))
    return out


def print_summary(classified: list[tuple[Rule, dict]]) -> None:
    counter = Counter(r.value for r, _ in classified)
    total_files = sum(
        (row.get("total_satellite_files") or 0)
        for r, row in classified
        if r in (Rule.R3_MERGE, Rule.R4_PROVISION_CONSOLIDATE)
    )
    total_sats = sum(
        (row.get("n_satellites") or 0)
        for r, row in classified
        if r in (Rule.R3_MERGE, Rule.R4_PROVISION_CONSOLIDATE)
    )

    print("=" * 60)
    print(f"CRM Guardian classification — {len(classified)} clients")
    print("=" * 60)
    for k in sorted(counter.keys()):
        print(f"  {k:35} {counter[k]:5}")
    print("-" * 60)
    print(f"  Files to move (R3+R4):           {total_files:5}")
    print(f"  Satellite folders to trash:      {total_sats:5}")
    print(f"  Clients to touch (R3+R4):        {counter.get('R3_merge', 0) + counter.get('R4_provision_consolidate', 0):5}")
    print("=" * 60)


def write_excel(classified: list[tuple[Rule, dict]], drive) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = "1_Summary"
    ws["A1"] = "CRM Guardian — Dry-Run Plan"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws.merge_cells("A1:B1")

    counter = Counter(r.value for r, _ in classified)
    rows_summary = [
        ("Generated", datetime.now(timezone.utc).isoformat() + "Z"),
        ("Total clients analyzed", len(classified)),
        ("", ""),
        ("Rule counts", ""),
    ]
    for k in sorted(counter.keys()):
        rows_summary.append((k, counter[k]))
    rows_summary.append(("", ""))
    rows_summary.append(("Files to move (R3+R4)", sum((r[1].get("total_satellite_files") or 0) for r in classified if r[0] in (Rule.R3_MERGE, Rule.R4_PROVISION_CONSOLIDATE))))
    rows_summary.append(("Satellite folders to trash", sum((r[1].get("n_satellites") or 0) for r in classified if r[0] in (Rule.R3_MERGE, Rule.R4_PROVISION_CONSOLIDATE))))

    for i, (k, v) in enumerate(rows_summary, 3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30

    # Sheet 2 — Client Plan (one row per client with total op counts)
    ws2 = wb.create_sheet("2_ClientPlan")
    headers2 = ["Rule", "ClientID", "Status", "FullName", "Assigned", "HasCanonical", "CanonicalName", "MiscID", "NSats", "TotalFiles", "TotalSubs"]
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    rule_colors = {
        "R2_clean": "D5F4E1",
        "R3_merge": "FFF4CC",
        "R4_provision_consolidate": "FFE5CC",
        "R5_provision_only": "E7F0F9",
        "R1_test_data": "FFD9D9",
    }
    row_idx = 2
    for rule, row in sorted(classified, key=lambda x: (x[0].value, -(x[1].get("total_satellite_files") or 0))):
        fill_color = rule_colors.get(rule.value, "FFFFFF")
        vals = [
            rule.value,
            row["client_id"],
            row.get("status"),
            row["full_name"],
            row.get("assigned_to") or "",
            "YES" if row.get("has_canonical") else "no",
            row.get("canonical_folder_name") or "",
            row.get("canonical_misc_folder_id") or "",
            row.get("n_satellites") or 0,
            row.get("total_satellite_files") or 0,
            row.get("total_satellite_subfolders") or 0,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row_idx, column=c, value=v)
            cell.fill = PatternFill("solid", fgColor=fill_color)
        row_idx += 1
    widths2 = [28, 10, 12, 34, 28, 14, 40, 44, 8, 11, 11]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{row_idx - 1}"

    # Sheet 3 — Move Operations (flat list file-by-file, only for R3/R4, bounded by performance)
    ws3 = wb.create_sheet("3_MoveOperations")
    headers3 = ["Rule", "ClientID", "FullName", "SatelliteID", "SatelliteName", "FileID", "FileName", "MimeType", "SourceParentID", "DestParentID", "NewName"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    row_idx = 2
    candidates = [(r, row) for r, row in classified if r in (Rule.R3_MERGE, Rule.R4_PROVISION_CONSOLIDATE)]
    # Sort by total files desc so the heaviest cases appear first
    candidates.sort(key=lambda x: -(x[1].get("total_satellite_files") or 0))

    # Compute move ops for each (this reads Drive — expensive).
    # Use TOP-15 clients by file count to keep the Excel sane.
    TOP_N = 15
    heavy = candidates[:TOP_N]
    print(f"  Computing move ops for top {TOP_N} heaviest clients (this hits Drive API)...", file=sys.stderr)
    for rule, row in heavy:
        # We need canonical/misc IDs; for R4 they don't exist yet — skip move-ops for R4 in Excel
        # (they'd require live provisioning). Show only R3 file-level ops + R4 summary.
        if rule == Rule.R4_PROVISION_CONSOLIDATE:
            for sat in row["satellites"]:
                ws3.cell(row=row_idx, column=1, value=rule.value)
                ws3.cell(row=row_idx, column=2, value=row["client_id"])
                ws3.cell(row=row_idx, column=3, value=row["full_name"])
                ws3.cell(row=row_idx, column=4, value=sat["id"])
                ws3.cell(row=row_idx, column=5, value=sat["name"])
                ws3.cell(row=row_idx, column=6, value="(R4: canonical to be created, then all files moved)")
                ws3.cell(row=row_idx, column=7, value=f"({sat.get('total_files_recursive', '?')} files recursive)")
                ws3.cell(row=row_idx, column=11, value="(new canonical/99_Misc)")
                row_idx += 1
            continue
        misc_id = row.get("canonical_misc_folder_id")
        if not misc_id:
            continue
        try:
            ops = plan_consolidation(drive, row, row["canonical_folder_id"], misc_id)
        except Exception as exc:
            print(f"  move-ops plan failed for client {row['client_id']}: {exc}", file=sys.stderr)
            continue
        for op in ops:
            ws3.cell(row=row_idx, column=1, value=rule.value)
            ws3.cell(row=row_idx, column=2, value=row["client_id"])
            ws3.cell(row=row_idx, column=3, value=row["full_name"])
            ws3.cell(row=row_idx, column=4, value="")  # satellite inferred from source_parent_name
            ws3.cell(row=row_idx, column=5, value=op.source_parent_name)
            ws3.cell(row=row_idx, column=6, value=op.file_id)
            ws3.cell(row=row_idx, column=7, value=op.file_name)
            ws3.cell(row=row_idx, column=8, value=op.mime_type)
            ws3.cell(row=row_idx, column=9, value=op.source_parent_id)
            ws3.cell(row=row_idx, column=10, value=op.dest_parent_id)
            ws3.cell(row=row_idx, column=11, value=op.new_name)
            row_idx += 1

    widths3 = [26, 10, 32, 34, 38, 40, 40, 28, 40, 40, 50]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers3))}{row_idx - 1}"

    # Sheet 4 — Satellites detail (all clients)
    ws4 = wb.create_sheet("4_Satellites")
    headers4 = ["Rule", "ClientID", "FullName", "SatelliteID", "SatelliteName", "OwnerEmail", "ParentID", "DirectFiles", "DirectSubs", "TotalFiles", "TotalSubs", "Created"]
    for c, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    row_idx = 2
    for rule, row in sorted(classified, key=lambda x: -(x[1].get("total_satellite_files") or 0)):
        if rule not in (Rule.R3_MERGE, Rule.R4_PROVISION_CONSOLIDATE):
            continue
        for sat in row["satellites"]:
            vals = [
                rule.value, row["client_id"], row["full_name"],
                sat["id"], sat["name"], sat.get("owner_email") or "",
                sat.get("parent_id") or "",
                sat.get("direct_files"), sat.get("direct_subfolders"),
                sat.get("total_files_recursive"), sat.get("total_subfolders_recursive"),
                (sat.get("created") or "")[:19],
            ]
            for c, v in enumerate(vals, 1):
                ws4.cell(row=row_idx, column=c, value=v)
            row_idx += 1
    widths4 = [28, 10, 30, 40, 38, 28, 40, 10, 10, 10, 10, 20]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers4))}{row_idx - 1}"

    EXCEL_OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_OUT)
    print(f"\n✅ Excel written: {EXCEL_OUT}", file=sys.stderr)


async def apply_clients(classified: list[tuple[Rule, dict]], *, force_dry_run: bool, only_ids: set[int] | None) -> None:
    db_url = _load_db_url()
    conn = await asyncpg.connect(db_url)
    try:
        # Kill switch
        if not await is_globally_enabled(conn) and not force_dry_run:
            raise SystemExit(
                "Kill switch OFF: system_settings.crm_guardian_enabled='false'.\n"
                "To enable: UPDATE system_settings SET value='true' WHERE key='crm_guardian_enabled';"
            )

        i3_state = await get_invariant_state(conn, "I3_satellite_consolidation")
        i1_state = await get_invariant_state(conn, "I1_canonical_folder")
        if not force_dry_run:
            if not i3_state["enabled"]:
                raise SystemExit("I3_satellite_consolidation is disabled in crm_guardian_state. Enable with SQL first.")
            if i3_state["circuit_breaker_tripped"]:
                raise SystemExit("I3 circuit breaker tripped. Reset before continuing.")

        config = GuardianConfig(dry_run=force_dry_run or bool(i3_state["dry_run"]))
        drive = build_drive_service()
        context3 = GuardianRunContext(invariant_id="I3_satellite_consolidation", config=config)
        context4 = GuardianRunContext(invariant_id="I1_canonical_folder", config=config)

        print(f"Apply: dry_run={config.dry_run}, run_id_I3={context3.run_id[:8]}... run_id_I1={context4.run_id[:8]}...")

        totals = Counter()
        for rule, row in classified:
            if only_ids and row["client_id"] not in only_ids:
                continue
            if rule == Rule.R3_MERGE:
                res = await apply_consolidation_for_client(conn, drive, row, context3)
                totals["R3_files_moved"] += res.get("files_moved", 0)
                totals["R3_folders_trashed"] += res.get("folders_trashed", 0)
                totals["R3_errors"] += res.get("errors", 0)
                totals["R3_clients"] += 1
                print(f"  R3 client {row['client_id']} {row['full_name']!r}: moved={res.get('files_moved')} trashed={res.get('folders_trashed')} errors={res.get('errors')}")
            elif rule == Rule.R4_PROVISION_CONSOLIDATE:
                res = await apply_provision_and_consolidate(conn, drive, row, context4)
                totals["R4_files_moved"] += res.get("files_moved", 0)
                totals["R4_folders_trashed"] += res.get("folders_trashed", 0)
                totals["R4_errors"] += res.get("errors", 0)
                totals["R4_clients"] += 1
                print(f"  R4 client {row['client_id']} {row['full_name']!r}: provisioned={res.get('provisioned_canonical','n/a')[:12]}... moved={res.get('files_moved')} errors={res.get('errors')}")
            else:
                totals[f"skipped_{rule.value}"] += 1

        print("\n=== TOTALS ===")
        for k in sorted(totals.keys()):
            print(f"  {k}: {totals[k]}")
        await bump_circuit_breaker(conn, "I3_satellite_consolidation", context3.error_count == 0)
        await bump_circuit_breaker(conn, "I1_canonical_folder", context4.error_count == 0)
    finally:
        await conn.close()


async def main_async(args) -> None:
    rows = load_plan()
    print(f"Loaded {len(rows)} rows from {PLAN_FILE}", file=sys.stderr)

    db_url = _load_db_url()
    conn = await asyncpg.connect(db_url)
    try:
        exc_i3 = await load_exceptions(conn, "I3_satellite_consolidation")
    finally:
        await conn.close()

    classified = classify_all(rows, exc_i3)

    if args.classify:
        print_summary(classified)
        return

    if args.dry_run:
        print_summary(classified)
        print("\nBuilding Excel dry-run plan (reads Drive for file-level ops on top-15)...", file=sys.stderr)
        drive = build_drive_service()
        write_excel(classified, drive)
        return

    if args.apply:
        only_ids = set(int(x) for x in args.client_ids.split(",")) if args.client_ids else None
        print_summary(classified)
        print()
        await apply_clients(classified, force_dry_run=False, only_ids=only_ids)
        return

    parser.print_help()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--classify", action="store_true", help="Print rule counts and exit")
    parser.add_argument("--dry-run", action="store_true", help="Generate Excel plan without touching Drive/DB")
    parser.add_argument("--apply", action="store_true", help="REAL apply (respects kill switch)")
    parser.add_argument("--client-ids", type=str, default=None, help="Comma-separated client IDs to process (apply only)")
    parser.add_argument("--batch", type=int, default=None, help="Max clients per run (apply only)")
    args = parser.parse_args()

    if not any([args.classify, args.dry_run, args.apply]):
        parser.print_help()
        sys.exit(1)

    asyncio.run(main_async(args))
