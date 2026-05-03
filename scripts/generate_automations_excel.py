#!/usr/bin/env python3
"""
Generate automations-inventory.xlsx from live system state.

Sources:
  - job_registry.json (sentinel-tracked jobs)
  - circuit_breakers.json (health state)
  - crontab -l (Pro + Air via SSH)
  - launchctl list (Pro + Air via SSH)
  - openclaw cron/jobs.json (Pro + Air via SSH)
  - MODEL_TOPOLOGY.json (model assignments)

Output: docs/automations-inventory.xlsx

Run:  python3 scripts/generate_automations_excel.py
Cron: alongside generate_automations_reference.py (nightly 23:15)
"""
import json
import os
import re
import subprocess
import socket
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
HOME = Path.home()
NUZANTARA_ROOT = Path(__file__).parent.parent
OUTPUT_FILE = NUZANTARA_ROOT / "docs" / "automations-inventory.xlsx"
REGISTRY_PATH = HOME / ".agent" / "decisions" / "job_registry.json"
CB_PATH = HOME / ".agent" / "decisions" / "circuit_breakers.json"
TOPOLOGY_PATH = NUZANTARA_ROOT / "MODEL_TOPOLOGY.json"
OPENCLAW_JOBS_PRO = HOME / ".openclaw" / "cron" / "jobs.json"

HOSTNAME = socket.gethostname()
IS_PRO = HOSTNAME == "Nuzantara"

# ── Colors ────────────────────────────────────────────────────────────────────
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
BLUE_FILL = PatternFill("solid", fgColor="BDD7EE")
GRAY_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


@dataclass
class Automation:
    name: str
    machine: str  # Pro / Air
    type: str  # cron / launchd / openclaw / daemon
    system: str = ""  # Garuda / Olympus / Cell / SEO / CRM / NLM / Infra / Ops / Intel
    schedule: str = ""
    schedule_human: str = ""
    script_path: str = ""
    description: str = ""
    uses_llm: str = ""  # model name or "—"
    llm_interface: str = ""  # ollama / claude-cli / gemini-cli / openclaw / —
    monitored_by: str = ""  # sentinel / — / manual
    is_critical: bool = False
    repair_scope: str = ""
    status: str = ""  # healthy / failed / terminal / unknown / skipped
    circuit_state: str = ""
    failures: int = 0
    last_run: str = ""
    last_error: str = ""
    produces: str = ""
    consumes: str = ""
    tools_called: str = ""
    apis_called: str = ""
    secrets_used: str = ""
    notes: str = ""


# ── System classification ────────────────────────────────────────────────────
# Maps automation name patterns to the system they belong to.
SYSTEM_RULES: list[tuple[str, str]] = [
    # Garuda — Knowledge Graph intelligence
    ("garuda", "Garuda"),
    ("gap-detector", "Garuda"),
    ("knowledge-graph", "Garuda"),
    ("matagaruda", "Garuda"),
    # Cell — Autonomous organism
    ("cell", "Cell"),
    # NLM — NotebookLM pipelines
    ("nlm", "NLM"),
    ("nb1", "NLM"), ("nb2", "NLM"), ("nb3", "NLM"), ("nb4", "NLM"),
    ("nb5", "NLM"), ("nb6", "NLM"), ("nb7", "NLM"), ("nb8", "NLM"),
    ("nb10", "NLM"), ("notebook", "NLM"),
    ("freshness", "NLM"), ("gap_scanner", "NLM"), ("multimodal", "NLM"),
    ("heartbeat_check", "NLM"), ("ops_briefing", "NLM"), ("persona_validate", "NLM"),
    ("peraturan_ingestion", "NLM"), ("peraturan-ingestion", "NLM"),
    ("db_nlm_sync", "NLM"), ("db-nlm-sync", "NLM"),
    ("yt_monitor", "NLM"), ("yt-monitor", "NLM"), ("deep-research", "NLM"),
    ("run_gap", "NLM"), ("gap_scanner", "NLM"),
    ("run_heartbeat", "NLM"), ("heartbeat_check", "NLM"),
    ("run_ops_briefing", "NLM"), ("ops_briefing", "NLM"),
    ("run_persona", "NLM"), ("persona_validate", "NLM"),
    ("run_nb", "NLM"), ("run_freshness", "NLM"),
    ("run_multimodal", "NLM"),
    # SEO — Search engine optimization
    ("seo", "SEO"), ("indexing", "SEO"), ("kbli", "SEO"),
    # CRM — Client relationship management
    ("client", "CRM"), ("compliance", "CRM"), ("practice", "CRM"),
    ("renewal", "CRM"), ("expiry", "CRM"), ("cashout", "CRM"),
    ("notifier", "CRM"), ("birthday", "CRM"), ("welcome", "CRM"),
    ("crm", "CRM"), ("daily-ops", "CRM"), ("weekly-review", "CRM"),
    # Garuda — Intelligence: scraping + KG + regulation radar
    ("intel", "Garuda"), ("scraper", "Garuda"), ("legal_radar", "Garuda"),
    ("war-room", "Garuda"), ("normativa", "Garuda"),
    # Sentinel — Self-healing & monitoring
    ("sentinel", "Sentinel"), ("dlq", "Sentinel"), ("circuit", "Sentinel"),
    ("zombie", "Sentinel"), ("disk-monitor", "Sentinel"),
    ("fly-health", "Sentinel"), ("fly-restart", "Sentinel"),
    ("cert-monitor", "Sentinel"), ("fly-cost", "Sentinel"),
    ("doctor", "Sentinel"), ("canary", "Sentinel"),
    ("ragas", "Sentinel"), ("guardian-ragas", "Sentinel"),
    ("guardian-redteam", "Sentinel"), ("job_health", "Sentinel"),
    ("automations-reference", "Sentinel"), ("automap", "Sentinel"),
    ("monitor", "Sentinel"), ("healthcheck", "Sentinel"),
    ("core-guardian", "Sentinel"), ("tech-orchestrator", "Sentinel"),
    ("vector-reindex", "Sentinel"),
    # Olympus — Database & infrastructure
    ("postgres", "Olympus"), ("pg-sync", "Olympus"), ("pg-backup", "Olympus"),
    ("fly-backup", "Olympus"), ("qdrant", "Olympus"),
    # Ops — Operational tools
    ("translate", "Ops"), ("overnight", "Ops"),
    ("openclaw", "Ops"), ("nuz-sync", "Ops"), ("syncthing", "Ops"),
    ("sync-damar", "Ops"), ("sync-memory", "Ops"), ("mos", "Ops"),
    ("memory", "Ops"), ("cache", "Ops"), ("cleanup", "Ops"),
    ("warm", "Ops"), ("ollama", "Ops"), ("redis", "Ops"),
    ("tunnel", "Ops"), ("claude-max", "Ops"), ("docker", "Ops"),
    ("nightly-sync", "Ops"), ("drive", "Ops"), ("state-bridge", "Ops"),
    ("webhook", "Ops"), ("poller", "Ops"), ("code-review", "Ops"),
    ("conversation-cleanup", "Ops"), ("conversation-trainer", "Ops"),
    ("dep-audit", "Ops"), ("dep_audit", "Ops"), ("coverage", "Ops"),
    ("learning-pipeline", "Ops"), ("codebase-audit", "Ops"),
    ("weekly-report", "Ops"), ("weekly-dep", "Ops"),
    ("t4-monitor", "Garuda"), ("t4_monitor", "Garuda"),
    # Sentinel additional
    ("auto_test", "Sentinel"), ("auto-test", "Sentinel"),
    ("auto_sentinel", "Sentinel"), ("auto-sentinel", "Sentinel"),
    ("auto_judgement", "Sentinel"), ("judgement-day", "Sentinel"),
    ("job-health", "Sentinel"), ("job_health", "Sentinel"),
    # Garuda additional
    ("source-enrichment", "Garuda"), ("legal-radar", "Garuda"),
    ("auto_kb_ingest", "Sentinel"), ("kb-ingest", "Sentinel"),
    # Ops additional
    ("cron-wrapper", "Ops"), ("cron_wrapper", "Ops"),
    ("heartbeat", "Ops"), ("pro_heartbeat", "Ops"),
]


def classify_system(name: str) -> str:
    """Classify an automation into a system based on name patterns."""
    name_lower = name.lower()
    for pattern, system in SYSTEM_RULES:
        # Match both underscore and dash variants
        if pattern in name_lower or pattern.replace("_", "-") in name_lower or pattern.replace("-", "_") in name_lower:
            return system
    return "—"


# ── Catalog ───────────────────────────────────────────────────────────────────
CATALOG_PATH = Path(__file__).parent / "automation_catalog.json"
AUTO_DISCOVERED_SECTION = "auto_discovered"


def load_catalog() -> dict:
    """Load human-verified automation catalog for enrichment."""
    try:
        return json.loads(CATALOG_PATH.read_text())
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def save_catalog(catalog: dict) -> None:
    """Atomically save catalog back to disk."""
    tmp = str(CATALOG_PATH) + ".tmp"
    Path(tmp).write_text(json.dumps(catalog, indent=2, ensure_ascii=False))
    Path(tmp).replace(CATALOG_PATH)


def _find_catalog_entry(name: str, catalog: dict) -> Optional[dict]:
    """Find an entry in any catalog section by name or basename."""
    for section in ("openclaw_pro", "openclaw_air", "launchagents",
                    "cron_scripts", "nlm_pipelines", "backend_services",
                    "github_actions", "claude_code_hooks", "home_scripts",
                    "air_cron_extras", AUTO_DISCOVERED_SECTION):
        entry = catalog.get(section, {}).get(name)
        if entry:
            return entry
    # Fallback: match by basename
    basename = os.path.basename(name) if "/" in name else name
    for section in ("cron_scripts", "nlm_pipelines", AUTO_DISCOVERED_SECTION):
        entry = catalog.get(section, {}).get(basename)
        if entry:
            return entry
    return None


def enrich_from_catalog(a: Automation, catalog: dict) -> None:
    """Enrich an Automation from the catalog. Catalog wins over auto-detected."""
    entry = _find_catalog_entry(a.name, catalog)
    if not entry:
        return

    if entry.get("description") and (not a.description or a.description.startswith("(empty")):
        a.description = entry["description"]
    if entry.get("produces") and not a.produces:
        a.produces = entry["produces"]
    if entry.get("consumes") and not a.consumes:
        a.consumes = entry["consumes"]
    if entry.get("uses_llm") and (not a.uses_llm or a.uses_llm == "—"):
        a.uses_llm = entry["uses_llm"]
    if entry.get("llm_interface") and (not a.llm_interface or a.llm_interface == "—"):
        a.llm_interface = entry["llm_interface"]
    if entry.get("notes"):
        a.notes = entry["notes"] if not a.notes else a.notes + "; " + entry["notes"]
    if entry.get("type"):
        a.type = entry["type"]
    if entry.get("monitored_by") and not a.monitored_by:
        a.monitored_by = entry["monitored_by"]
    if entry.get("schedule") and (not a.schedule_human or a.schedule_human in ("at boot", "daemon", "")):
        a.schedule_human = entry["schedule"]
    # New enrichment fields: tools, APIs, secrets
    if entry.get("tools_called") and not a.tools_called:
        tc = entry["tools_called"]
        a.tools_called = ", ".join(tc) if isinstance(tc, list) else str(tc)
    if entry.get("apis_called") and not a.apis_called:
        ac = entry["apis_called"]
        a.apis_called = ", ".join(ac) if isinstance(ac, list) else str(ac)
    if entry.get("secrets_used") and not a.secrets_used:
        su = entry["secrets_used"]
        a.secrets_used = ", ".join(su) if isinstance(su, list) else str(su)


def _read_script_header(script_path: str) -> str:
    """Read the first comment block from a script file (up to 5 lines)."""
    try:
        path = os.path.expanduser(script_path)
        if not os.path.exists(path):
            return ""
        lines = Path(path).read_text(errors="ignore").splitlines()[:10]
        comments = []
        for line in lines:
            stripped = line.strip()
            if stripped.startswith("#!"):
                continue  # skip shebang
            if stripped.startswith("#"):
                comments.append(stripped.lstrip("# ").strip())
            elif stripped.startswith('"""') or stripped.startswith("'''"):
                # Python docstring
                comments.append(stripped.strip("\"'").strip())
            elif comments:
                break  # stop at first non-comment line after we found some
        return " ".join(comments)[:200]
    except Exception:
        return ""


def _llm_describe_script(script_path: str, name: str) -> Optional[str]:
    """Use claude --print to generate a one-line description of an unknown script."""
    try:
        header = _read_script_header(script_path)
        if not header:
            path = os.path.expanduser(script_path)
            if os.path.exists(path):
                header = Path(path).read_text(errors="ignore")[:500]
            else:
                return None

        prompt = (
            f"Describe this automation script in ONE sentence (max 120 chars). "
            f"Script name: {name}\n"
            f"Content:\n{header}\n\n"
            f"Reply with ONLY the description, no quotes, no prefix."
        )
        proc = subprocess.run(
            ["claude", "--print", prompt],
            capture_output=True, text=True, timeout=20,
        )
        if proc.returncode == 0 and proc.stdout.strip():
            return proc.stdout.strip()[:200]
    except (subprocess.TimeoutExpired, Exception):
        pass
    return None


def auto_discover_unknown(automations: list, catalog: dict) -> int:
    """
    For automations not in catalog: try header parsing, then LLM fallback.
    Writes new entries to catalog[AUTO_DISCOVERED_SECTION] and saves to disk.
    Returns count of newly discovered entries.
    """
    discovered = 0
    section = catalog.setdefault(AUTO_DISCOVERED_SECTION, {})

    for a in automations:
        # Skip if already in catalog
        if _find_catalog_entry(a.name, catalog):
            continue
        # Skip if already has a description from OpenClaw payload
        if a.description and not a.description.startswith("(empty"):
            continue
        # Skip inline commands (not script files)
        if not a.script_path:
            continue

        # Step 1: try reading script header
        header_desc = _read_script_header(a.script_path)

        if header_desc and len(header_desc) > 15:
            desc = header_desc
        else:
            # Step 2: LLM fallback (only if claude CLI is available)
            desc = _llm_describe_script(a.script_path, a.name)

        if desc:
            entry = {"description": desc, "_source": "auto_discovered"}
            section[a.name] = entry
            a.description = desc
            discovered += 1

    if discovered > 0:
        save_catalog(catalog)

    return discovered


# ── Data collection ──────────────────────────────────────────────────────────

def load_json(path: Path) -> dict:
    try:
        return json.loads(path.read_text())
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def ssh_cmd(cmd: str, timeout: int = 10) -> str:
    try:
        r = subprocess.run(
            ["ssh", "-o", "ConnectTimeout=5", "-o", "BatchMode=yes", "air", cmd],
            capture_output=True, text=True, timeout=timeout,
        )
        return r.stdout
    except Exception:
        return ""


def parse_cron_schedule(line: str) -> tuple[str, str, str]:
    """Extract (schedule, command, human-readable) from a crontab line."""
    line = line.strip()
    if not line or line.startswith("#") or line.startswith("PATH=") or line.startswith("MAILTO="):
        return "", "", ""
    parts = line.split(None, 5)
    if len(parts) < 6:
        return "", "", ""
    sched = " ".join(parts[:5])
    cmd = parts[5]
    # Human-readable
    human = cron_to_human(sched)
    return sched, cmd, human


def cron_to_human(expr: str) -> str:
    """Simple cron expression to human-readable."""
    parts = expr.split()
    if len(parts) != 5:
        return expr
    minute, hour, dom, month, dow = parts
    days = {"0": "Dom", "1": "Lun", "2": "Mar", "3": "Mer", "4": "Gio", "5": "Ven", "6": "Sab"}

    result = []
    if minute.startswith("*/"):
        result.append(f"ogni {minute[2:]}min")
    elif hour.startswith("*/"):
        result.append(f"ogni {hour[2:]}h :{minute}")
    elif hour != "*":
        result.append(f"{hour}:{minute.zfill(2)} WITA")
    else:
        result.append(f"ogni ora :{minute}")

    if dow != "*":
        dow_names = [days.get(d.strip(), d) for d in dow.split(",")]
        result.append(" ".join(dow_names))
    if dom != "*":
        result.append(f"giorno {dom}")

    return " ".join(result)


def detect_llm_usage(script_path: str) -> tuple[str, str]:
    """Check if a script uses LLM. Returns (model, interface)."""
    if not script_path or not os.path.exists(script_path):
        return "", ""
    try:
        content = Path(script_path).read_text(errors="ignore")[:5000]
    except Exception:
        return "", ""

    if "ollama" in content.lower() or "11434" in content:
        model_match = re.search(
            r'(qwen[0-9.]*:[0-9a-z]+|gemma[0-9]*:[0-9a-z]+|deepseek[^"\'  ]+)',
            content,
        )
        model = model_match.group() if model_match else "ollama (unknown)"
        return model, "Ollama API"
    if "claude" in content.lower() and "print" in content.lower():
        return "claude CLI", "claude --print"
    if "gemini" in content.lower():
        return "gemini CLI", "gemini --print"
    return "", ""


def extract_script_name(cmd: str) -> str:
    """Extract meaningful script name from a crontab command."""
    # Find the last .sh or .py file in the command
    matches = re.findall(r'[\w/.-]+\.(?:sh|py)', cmd)
    if matches:
        return os.path.basename(matches[-1])
    # curl commands
    if "curl" in cmd and "nuzantara-rag" in cmd:
        endpoint = re.search(r'/api/[^\s"]+', cmd)
        if endpoint:
            return f"API: {endpoint.group()}"
    return cmd[:60]


def collect_crontab(machine: str) -> list[Automation]:
    """Collect crontab entries."""
    if machine == "Pro":
        result = subprocess.run(["crontab", "-l"], capture_output=True, text=True)
        lines = result.stdout.splitlines()
    else:
        lines = ssh_cmd("crontab -l").splitlines()

    automations = []
    for line in lines:
        sched, cmd, human = parse_cron_schedule(line)
        if not sched:
            continue
        script = extract_script_name(cmd)
        # Detect LLM
        script_paths = re.findall(r'(/[\w/.-]+\.(?:sh|py))', cmd)
        model, interface = "", ""
        for sp in script_paths:
            real_path = sp
            if machine == "Air":
                # Can't read Air files directly, mark for manual check
                pass
            else:
                model, interface = detect_llm_usage(real_path)
                if model:
                    break

        # Detect API calls
        if "curl" in cmd and "nuzantara-rag.fly.dev" in cmd:
            interface = "Fly.io API"

        # Detect NLM
        if "nlm" in cmd.lower() or "notebooklm" in cmd.lower() or "sync-memory-to-nlm" in cmd:
            interface = "NLM CLI (cloud)"

        a = Automation(
            name=script,
            machine=machine,
            type="cron",
            schedule=sched,
            schedule_human=human,
            script_path=script_paths[0] if script_paths else "",
            uses_llm=model if model else "—",
            llm_interface=interface if interface else "—",
            monitored_by="—",
        )
        automations.append(a)
    return automations


def collect_launchagents(machine: str) -> list[Automation]:
    """Collect LaunchAgent entries."""
    if machine == "Pro":
        plist_dir = HOME / "Library" / "LaunchAgents"
        plists = list(plist_dir.glob("*.plist"))
    else:
        raw = ssh_cmd("ls ~/Library/LaunchAgents/*.plist 2>/dev/null")
        plists = [Path(p.strip()) for p in raw.splitlines() if p.strip()]

    automations = []
    for plist in plists:
        label = plist.stem
        # Skip system/google/adobe
        if any(x in label for x in ["google", "adobe", "openai.atlas"]):
            continue

        if machine == "Pro":
            # Read plist to get program
            try:
                raw_plist = subprocess.run(
                    ["defaults", "read", str(plist), "ProgramArguments"],
                    capture_output=True, text=True,
                )
                prog = raw_plist.stdout.strip()
            except Exception:
                prog = ""

            # Check RunAtLoad / KeepAlive / StartInterval
            try:
                keep = subprocess.run(
                    ["defaults", "read", str(plist), "KeepAlive"],
                    capture_output=True, text=True,
                ).stdout.strip()
            except Exception:
                keep = ""

            try:
                interval = subprocess.run(
                    ["defaults", "read", str(plist), "StartInterval"],
                    capture_output=True, text=True,
                ).stdout.strip()
            except Exception:
                interval = ""

            is_daemon = keep == "1" or "true" in keep.lower()
            sched = f"every {interval}s" if interval and interval != "0" else ("daemon" if is_daemon else "at boot")
        else:
            prog = ""
            sched = ""
            is_daemon = False

        a_type = "daemon" if is_daemon else "launchd"

        a = Automation(
            name=label,
            machine=machine,
            type=a_type,
            schedule_human=sched,
            uses_llm="—",
            llm_interface="—",
            monitored_by="—",
        )
        automations.append(a)
    return automations


def collect_openclaw(machine: str) -> list[Automation]:
    """Collect OpenClaw cron jobs."""
    if machine == "Pro":
        data = load_json(OPENCLAW_JOBS_PRO)
    else:
        raw = ssh_cmd("cat ~/.openclaw/cron/jobs.json 2>/dev/null")
        try:
            data = json.loads(raw)
        except (json.JSONDecodeError, Exception):
            data = {}

    topology = load_json(TOPOLOGY_PATH)
    warm_model = topology.get("nodes", {}).get(
        "pro" if machine == "Pro" else "air", {}
    ).get("warm_model", "?")

    automations = []
    for j in data.get("jobs", []):
        name = j.get("name", "?")
        agent = j.get("agentId", "main")
        state = j.get("state", {})
        payload = j.get("payload", {})
        msg = payload.get("message", "")
        sched = j.get("schedule", {})

        # Parse schedule
        kind = sched.get("kind", "?")
        if kind == "cron":
            expr = sched.get("expr", "?")
            human = cron_to_human(expr)
        elif kind == "every":
            ms = sched.get("everyMs", 0)
            if ms >= 3600000:
                human = f"ogni {ms // 3600000}h"
            elif ms >= 60000:
                human = f"ogni {ms // 60000}min"
            else:
                human = f"ogni {ms}ms"
        else:
            human = kind

        # Status
        ce = state.get("consecutiveErrors", 0)
        ls = state.get("lastRunStatus", "?")
        if ls == "ok":
            status = "healthy"
        elif ls == "skipped":
            status = "skipped"
        elif ce > 0:
            status = "failed"
        else:
            status = "unknown"

        # Description from message
        desc = msg[:200] if isinstance(msg, str) else str(msg)[:200]

        # Is it empty/skipped?
        if not msg or (isinstance(msg, str) and len(msg.strip()) < 10):
            status = "skipped"
            desc = "(empty task — placeholder)"

        a = Automation(
            name=name,
            machine=machine,
            type="openclaw",
            schedule_human=human,
            description=desc.replace("\n", " "),
            uses_llm=warm_model,
            llm_interface=f"OpenClaw ({agent})",
            monitored_by="Sentinel",
            status=status,
            failures=ce,
            last_run=datetime.fromtimestamp(
                state.get("lastRunAtMs", 0) / 1000, tz=timezone.utc
            ).strftime("%Y-%m-%d %H:%M") if state.get("lastRunAtMs") else "",
            last_error=state.get("lastError", "")[:100] if ce > 0 else "",
        )
        automations.append(a)
    return automations


def _get_launchagent_status(label: str, machine: str) -> tuple[str, str]:
    """Get LaunchAgent status from launchctl. Returns (status, exit_code_str)."""
    try:
        if machine == "Pro":
            r = subprocess.run(
                ["launchctl", "list", label],
                capture_output=True, text=True, timeout=5,
            )
        else:
            r = subprocess.run(
                ["ssh", "-o", "ConnectTimeout=3", "-o", "BatchMode=yes", "air",
                 f"launchctl list {label}"],
                capture_output=True, text=True, timeout=10,
            )
        if r.returncode != 0:
            return "not loaded", ""
        # Parse LastExitStatus
        for line in r.stdout.splitlines():
            if "LastExitStatus" in line:
                code = line.split("=")[-1].strip().rstrip(";").strip()
                if code == "0":
                    return "healthy", "0"
                return "failed", code
            if "PID" in line and "=" in line:
                return "running", ""
    except Exception:
        pass
    return "unknown", ""


def enrich_from_registry(automations: list[Automation]) -> None:
    """Enrich automations with registry + circuit breaker + launchctl data."""
    registry = load_json(REGISTRY_PATH).get("jobs", {})
    cbs = load_json(CB_PATH)

    for a in automations:
        # Match by name (registry uses underscores, openclaw uses dashes)
        reg_name = a.name.replace("-", "_")
        reg = registry.get(reg_name, {})
        cb = cbs.get(reg_name, {})

        if reg:
            a.is_critical = reg.get("critical", False)
            a.repair_scope = reg.get("repair_scope", "")
            a.monitored_by = "Sentinel"
            if not a.description:
                a.description = reg.get("_note", "")

        if cb:
            a.circuit_state = cb.get("state", "")
            a.failures = max(a.failures, cb.get("failures", 0))
            if a.circuit_state == "OPEN":
                a.status = "failed"
            elif a.circuit_state == "TERMINAL":
                a.status = "terminal"
            elif a.circuit_state == "CLOSED" and not a.status:
                a.status = "healthy"

        # Fill defaults for non-sentinel-tracked automations
        if not a.monitored_by:
            if a.type in ("daemon", "launchd"):
                a.monitored_by = "launchd"
            elif a.type == "cron":
                a.monitored_by = "cron"
            elif a.type == "openclaw":
                a.monitored_by = "OpenClaw"

        # Get LaunchAgent status from launchctl if not already set
        if a.type in ("daemon", "launchd") and not a.status:
            status, _ = _get_launchagent_status(a.name, a.machine)
            a.status = status

        # Default status for cron jobs (if no circuit breaker data)
        if a.type == "cron" and not a.status:
            a.status = "active"


def status_fill(status: str) -> PatternFill:
    if status in ("healthy", "ok", "running", "active"):
        return GREEN_FILL
    if status in ("failed", "terminal", "not loaded"):
        return RED_FILL
    if status in ("skipped", "unknown"):
        return GRAY_FILL
    if status in ("warning",):
        return YELLOW_FILL
    return PatternFill()


# ── Excel generation ─────────────────────────────────────────────────────────

COLUMNS = [
    ("Nome", 30),
    ("Sistema", 10),
    ("Macchina", 8),
    ("Tipo", 10),
    ("Schedule", 25),
    ("Descrizione", 50),
    ("Usa LLM", 15),
    ("Interfaccia LLM", 18),
    ("Monitorato da", 12),
    ("Critico", 8),
    ("Repair Scope", 12),
    ("Stato", 10),
    ("Circuit Breaker", 12),
    ("Errori", 7),
    ("Ultimo Run", 16),
    ("Ultimo Errore", 40),
    ("Produce", 30),
    ("Consuma", 30),
    ("Tools Chiamati", 40),
    ("API Chiamate", 35),
    ("Secrets", 25),
    ("Note", 35),
]

# System colors
SYSTEM_COLORS = {
    "Garuda": PatternFill("solid", fgColor="E2EFDA"),    # sage green
    "Cell": PatternFill("solid", fgColor="FCE4D6"),      # light orange
    "NLM": PatternFill("solid", fgColor="D6E4F0"),       # light blue
    "SEO": PatternFill("solid", fgColor="FFF2CC"),       # light yellow
    "CRM": PatternFill("solid", fgColor="E2D9F3"),       # light purple
    "Sentinel": PatternFill("solid", fgColor="D5E8D4"),  # mint green
    "Olympus": PatternFill("solid", fgColor="DAE8FC"),   # steel blue
    "Ops": PatternFill("solid", fgColor="F5F5F5"),       # light gray
}


def write_sheet(wb: Workbook, sheet_name: str, automations: list[Automation]) -> None:
    ws = wb.create_sheet(title=sheet_name)

    # Headers
    for col_idx, (title, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze top row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Sort: critical first, then by type, then by name
    sorted_autos = sorted(
        automations,
        key=lambda a: (
            0 if a.status in ("failed", "terminal") else 1,
            0 if a.is_critical else 1,
            {"openclaw": 0, "daemon": 1, "launchd": 2, "cron": 3}.get(a.type, 4),
            a.name,
        ),
    )

    for row_idx, a in enumerate(sorted_autos, 2):
        values = [
            a.name,
            a.system or "—",
            a.machine,
            a.type,
            a.schedule_human,
            a.description,
            a.uses_llm,
            a.llm_interface,
            a.monitored_by or "—",
            "SI" if a.is_critical else "—",
            a.repair_scope or "—",
            a.status or "—",
            a.circuit_state or "—",
            a.failures if a.failures else "",
            a.last_run,
            a.last_error,
            a.produces or "—",
            a.consumes or "—",
            a.tools_called or "—",
            a.apis_called or "—",
            a.secrets_used or "—",
            a.notes,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGN
            cell.border = THIN_BORDER

        # Color Sistema column (col 2)
        sys_cell = ws.cell(row=row_idx, column=2)
        sys_fill = SYSTEM_COLORS.get(a.system)
        if sys_fill:
            sys_cell.fill = sys_fill
            sys_cell.font = Font(bold=True, size=10)

        # Color status column (col 12)
        status_cell = ws.cell(row=row_idx, column=12)
        fill = status_fill(a.status)
        if fill.fgColor:
            status_cell.fill = fill

        # Color circuit breaker column (col 13)
        cb_cell = ws.cell(row=row_idx, column=13)
        if a.circuit_state == "OPEN":
            cb_cell.fill = RED_FILL
        elif a.circuit_state == "CLOSED":
            cb_cell.fill = GREEN_FILL
        elif a.circuit_state == "HALF_OPEN":
            cb_cell.fill = YELLOW_FILL

        # Color critical (col 10)
        crit_cell = ws.cell(row=row_idx, column=10)
        if a.is_critical:
            crit_cell.fill = BLUE_FILL
            crit_cell.font = Font(bold=True, size=10)

    # Row height
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 30


def write_summary_sheet(wb: Workbook, all_autos: list[Automation]) -> None:
    ws = wb.create_sheet(title="Summary", index=0)

    ws["A1"] = "NUZANTARA — Automations Inventory"
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} WITA"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    # Stats
    row = 4
    stats = [
        ("Totale automazioni", len(all_autos)),
        ("", ""),
        ("Pro", sum(1 for a in all_autos if a.machine == "Pro")),
        ("Air", sum(1 for a in all_autos if a.machine == "Air")),
        ("", ""),
        ("OpenClaw cron", sum(1 for a in all_autos if a.type == "openclaw")),
        ("Crontab", sum(1 for a in all_autos if a.type == "cron")),
        ("LaunchAgent/Daemon", sum(1 for a in all_autos if a.type in ("launchd", "daemon"))),
        ("", ""),
        ("Usano LLM locale", sum(1 for a in all_autos if a.uses_llm != "—" and a.uses_llm)),
        ("Usano NLM (cloud)", sum(1 for a in all_autos if "NLM" in a.llm_interface)),
        ("Usano API Fly.io", sum(1 for a in all_autos if "API" in a.llm_interface)),
        ("Shell puro", sum(1 for a in all_autos if a.llm_interface == "—")),
        ("", ""),
        ("Healthy", sum(1 for a in all_autos if a.status == "healthy")),
        ("Failed", sum(1 for a in all_autos if a.status == "failed")),
        ("Skipped", sum(1 for a in all_autos if a.status == "skipped")),
        ("Terminal", sum(1 for a in all_autos if a.status == "terminal")),
        ("Unknown", sum(1 for a in all_autos if a.status in ("unknown", "", "—"))),
        ("", ""),
        ("Critici", sum(1 for a in all_autos if a.is_critical)),
        ("Monitorati da Sentinel", sum(1 for a in all_autos if a.monitored_by == "Sentinel")),
        ("Circuit OPEN", sum(1 for a in all_autos if a.circuit_state == "OPEN")),
        ("Circuit CLOSED", sum(1 for a in all_autos if a.circuit_state == "CLOSED")),
    ]

    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = Font(bold=bool(label), size=11)
        if value != "":
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.font = Font(size=11)
            val_cell.alignment = Alignment(horizontal="right")
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10

    # Model topology summary
    row += 2
    ws.cell(row=row, column=1, value="Model Topology").font = Font(bold=True, size=13, color="1F4E79")
    row += 1
    topology = load_json(TOPOLOGY_PATH)
    for node_id, node in topology.get("nodes", {}).items():
        ws.cell(row=row, column=1, value=f"{node_id.upper()} ({node.get('hostname', '?')})").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value=f"  Warm model: {node.get('warm_model', '?')}")
        ws.cell(row=row, column=2, value=f"RAM: {node.get('ram_gb', '?')}GB")
        row += 1
        ws.cell(row=row, column=1, value=f"  Context: {node.get('warm_ctx', '?')} tokens")
        ws.cell(row=row, column=2, value=f"MLX: {'SI' if node.get('mlx_enabled') else 'NO'}")
        row += 1
        row += 1


def main() -> None:
    print("Collecting automations...")

    all_automations: list[Automation] = []

    # Pro
    print("  Pro crontab...")
    all_automations.extend(collect_crontab("Pro"))
    print(f"    {sum(1 for a in all_automations if a.machine == 'Pro' and a.type == 'cron')} cron entries")

    print("  Pro LaunchAgents...")
    all_automations.extend(collect_launchagents("Pro"))
    print(f"    {sum(1 for a in all_automations if a.machine == 'Pro' and a.type in ('launchd', 'daemon'))} launchagents")

    print("  Pro OpenClaw...")
    all_automations.extend(collect_openclaw("Pro"))
    print(f"    {sum(1 for a in all_automations if a.machine == 'Pro' and a.type == 'openclaw')} openclaw jobs")

    # Air
    print("  Air crontab...")
    all_automations.extend(collect_crontab("Air"))
    print(f"    {sum(1 for a in all_automations if a.machine == 'Air' and a.type == 'cron')} cron entries")

    print("  Air LaunchAgents...")
    all_automations.extend(collect_launchagents("Air"))
    print(f"    {sum(1 for a in all_automations if a.machine == 'Air' and a.type in ('launchd', 'daemon'))} launchagents")

    print("  Air OpenClaw...")
    all_automations.extend(collect_openclaw("Air"))
    print(f"    {sum(1 for a in all_automations if a.machine == 'Air' and a.type == 'openclaw')} openclaw jobs")

    # Enrich from registry + circuit breakers
    print("  Enriching from registry + circuit breakers...")
    enrich_from_registry(all_automations)

    # Enrich from human-verified catalog
    print("  Enriching from automation catalog...")
    catalog = load_catalog()
    enriched = 0
    for a in all_automations:
        before_desc = a.description
        enrich_from_catalog(a, catalog)
        if a.description != before_desc:
            enriched += 1
    print(f"    {enriched} automations enriched from catalog")

    # Classify systems
    print("  Classifying systems...")
    for a in all_automations:
        a.system = classify_system(a.name)
    systems = {}
    for a in all_automations:
        s = a.system or "—"
        systems[s] = systems.get(s, 0) + 1
    print(f"    {', '.join(f'{s}:{c}' for s, c in sorted(systems.items()))}")

    # Fill defaults for monitored_by where still empty
    for a in all_automations:
        if not a.monitored_by or a.monitored_by == "—":
            if a.type in ("daemon", "launchd"):
                a.monitored_by = "launchd"
            elif a.type == "cron":
                a.monitored_by = "cron"
            elif a.type == "openclaw":
                a.monitored_by = "OpenClaw"

    # Auto-discover unknown automations (header parse + LLM fallback)
    print("  Auto-discovering unknown automations...")
    discovered = auto_discover_unknown(all_automations, catalog)
    print(f"    {discovered} new automations auto-described (saved to catalog)")

    # Generate Excel
    print(f"\nGenerating {OUTPUT_FILE}...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Summary first
    write_summary_sheet(wb, all_automations)

    # Pro sheet
    pro = [a for a in all_automations if a.machine == "Pro"]
    write_sheet(wb, "Pro", pro)

    # Air sheet
    air = [a for a in all_automations if a.machine == "Air"]
    write_sheet(wb, "Air", air)

    # Backend services (from catalog)
    backend_services = []
    for name, entry in catalog.get("backend_services", {}).items():
        if not isinstance(entry, dict):
            continue
        a = Automation(
            name=name,
            machine=entry.get("machine", "Fly.io"),
            type=entry.get("type", "backend-loop"),
            system=entry.get("system", ""),
            schedule_human=entry.get("schedule", ""),
            description=entry.get("description", ""),
            uses_llm=entry.get("uses_llm", "—"),
            llm_interface=entry.get("llm_interface", "—"),
            monitored_by=entry.get("monitored_by", "Health Monitor"),
            status="running",
            produces=entry.get("produces", ""),
            consumes=entry.get("consumes", ""),
            tools_called=", ".join(entry["tools_called"]) if isinstance(entry.get("tools_called"), list) else entry.get("tools_called", ""),
            apis_called=", ".join(entry["apis_called"]) if isinstance(entry.get("apis_called"), list) else entry.get("apis_called", ""),
            secrets_used=", ".join(entry["secrets_used"]) if isinstance(entry.get("secrets_used"), list) else entry.get("secrets_used", ""),
            notes=entry.get("notes", ""),
        )
        backend_services.append(a)
    if backend_services:
        write_sheet(wb, "Backend", backend_services)

    # Extra catalog sections: GitHub Actions, Claude Code Hooks, Home Scripts, Air Cron Extras
    extra_sections = [
        ("github_actions", "GitHub Actions", "GitHub", "ci-workflow"),
        ("claude_code_hooks", "Hooks", "Pro", "hook"),
        ("home_scripts", "Home Scripts", "Pro", "shell"),
        ("air_cron_extras", "Air Extras", "Air", "cron"),
        ("mata_garuda_pipeline", "Mata Garuda", "Pro", "pipeline"),
        ("intel_scraper_pipeline", "Intel Scraper", "Pro", "pipeline"),
    ]
    all_extras = []
    for section_key, sheet_name, default_machine, default_type in extra_sections:
        items = []
        for name, entry in catalog.get(section_key, {}).items():
            if not isinstance(entry, dict):
                continue
            a = Automation(
                name=name,
                machine=entry.get("machine", default_machine),
                type=entry.get("type", default_type),
                system=entry.get("system", ""),
                schedule_human=entry.get("schedule", ""),
                description=entry.get("description", ""),
                uses_llm=entry.get("uses_llm", "—"),
                llm_interface=entry.get("llm_interface", "—"),
                monitored_by=entry.get("monitored_by", "—"),
                status=entry.get("status", "running"),
                produces=entry.get("produces", ""),
                consumes=entry.get("consumes", ""),
                tools_called=", ".join(entry["tools_called"]) if isinstance(entry.get("tools_called"), list) else entry.get("tools_called", ""),
                apis_called=", ".join(entry["apis_called"]) if isinstance(entry.get("apis_called"), list) else entry.get("apis_called", ""),
                secrets_used=", ".join(entry["secrets_used"]) if isinstance(entry.get("secrets_used"), list) else entry.get("secrets_used", ""),
                notes=entry.get("notes", ""),
            )
            items.append(a)
        if items:
            write_sheet(wb, sheet_name, items)
            all_extras.extend(items)

    # All combined (including backend + extras)
    all_with_backend = all_automations + backend_services + all_extras
    write_sheet(wb, "All", all_with_backend)

    wb.save(str(OUTPUT_FILE))
    total = len(all_with_backend)
    print(f"\nDone! {total} automations → {OUTPUT_FILE}")
    print(f"  Pro: {len(pro)} | Air: {len(air)} | Backend: {len(backend_services)} | Extras: {len(all_extras)}")


if __name__ == "__main__":
    main()
