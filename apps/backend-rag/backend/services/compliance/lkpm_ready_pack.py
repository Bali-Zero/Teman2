"""
LKPM Ready Pack Generator — HTML output formatted for OSS copy-paste
                           + PDF/XLSX builder + Drive/email orchestrator.

All values are deterministic. Same input = same output, always.
Narrative uses pre-written templates with fill-in-the-blank, NOT AI generation.
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from html import escape as html_escape
from typing import TYPE_CHECKING, Any

import httpx

if TYPE_CHECKING:
    import asyncpg

from backend.app.models.lkpm import LKPMReadyPack, ValidationSeverity

logger = logging.getLogger(__name__)

# ── Module-level persistent httpx client (Golden Rule #10) ────────────────
_brevo_client: httpx.AsyncClient | None = None


def _get_brevo_client() -> httpx.AsyncClient:
    """Lazy-init persistent httpx client for Brevo (Golden Rule #10)."""
    global _brevo_client  # noqa: PLW0603 — singleton by design
    if _brevo_client is None or _brevo_client.is_closed:
        _brevo_client = httpx.AsyncClient(timeout=15.0)
    return _brevo_client


async def close_brevo_client() -> None:
    """Call from FastAPI lifespan shutdown."""
    global _brevo_client  # noqa: PLW0603
    if _brevo_client is not None and not _brevo_client.is_closed:
        await _brevo_client.aclose()
    _brevo_client = None


# Pre-written obstacle templates in Bahasa Indonesia (NO AI generation)
OBSTACLE_TEMPLATES = {
    "covid_impact": (
        "Dampak pandemi COVID-19 masih mempengaruhi operasional perusahaan, "
        "termasuk keterlambatan pengiriman peralatan dan pembatasan perjalanan "
        "tenaga kerja asing."
    ),
    "permit_delay": (
        "Proses perizinan yang memerlukan waktu lebih lama dari yang direncanakan, "
        "termasuk pengurusan izin {permit_type} yang masih dalam proses."
    ),
    "supply_chain": (
        "Gangguan rantai pasokan global menyebabkan keterlambatan pengadaan "
        "peralatan dan material yang diperlukan untuk realisasi investasi."
    ),
    "construction_delay": ("Proses pembangunan/renovasi mengalami keterlambatan akibat {reason}."),
    "market_conditions": (
        "Kondisi pasar yang belum stabil mempengaruhi rencana ekspansi "
        "dan realisasi investasi perusahaan."
    ),
    "no_obstacles": ("Tidak ada hambatan signifikan dalam realisasi investasi pada periode ini."),
}


def format_idr(amount: int) -> str:
    """Format amount in IDR with titik separator (Indonesian format)."""
    if amount == 0:
        return "0"
    formatted = f"{abs(amount):,}".replace(",", ".")
    return f"-{formatted}" if amount < 0 else formatted


def generate_ready_pack_html(pack: LKPMReadyPack) -> str:
    """Generate HTML Ready Pack for copy-paste into OSS."""
    logger.info(f"Generating Ready Pack HTML for draft {pack.draft_id}")

    # Validation status indicator
    validation_html = ""
    if pack.validation_summary:
        for alert in pack.validation_summary.alerts:
            color = {
                ValidationSeverity.GREEN: "#28a745",
                ValidationSeverity.YELLOW: "#ffc107",
                ValidationSeverity.RED: "#dc3545",
            }.get(alert.severity, "#6c757d")
            icon = {
                ValidationSeverity.GREEN: "&#10004;",
                ValidationSeverity.YELLOW: "&#9888;",
                ValidationSeverity.RED: "&#10006;",
            }.get(alert.severity, "&#8226;")
            validation_html += (
                f'<div style="color:{color};margin:2px 0;">{icon} {alert.message}</div>\n'
            )

    # Escape user-supplied strings to prevent XSS
    company = html_escape(pack.company_name)
    npwp = html_escape(pack.npwp or "—")
    nib = html_escape(pack.nib or "—")
    obstacles = html_escape(pack.narrative_obstacles or OBSTACLE_TEMPLATES["no_obstacles"])
    plans = html_escape(
        pack.narrative_plans
        or "Perusahaan akan melanjutkan realisasi investasi sesuai rencana yang telah disetujui.",
    )

    return f"""<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <title>LKPM Ready Pack — {company} — {pack.quarter} {pack.year}</title>
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 20px; color: #333; }}
        h1 {{ color: #1a5276; border-bottom: 2px solid #1a5276; padding-bottom: 8px; }}
        h2 {{ color: #2c3e50; margin-top: 24px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 12px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: left; }}
        th {{ background-color: #f8f9fa; font-weight: 600; }}
        td.amount {{ text-align: right; font-family: monospace; }}
        .header-info {{ background: #f8f9fa; padding: 16px; border-radius: 8px; margin-bottom: 20px; }}
        .header-info p {{ margin: 4px 0; }}
        .validation {{ background: #fff3cd; padding: 12px; border-radius: 6px; margin: 12px 0; }}
        .copy-hint {{ color: #6c757d; font-size: 0.85em; font-style: italic; }}
        .total-row {{ font-weight: bold; background-color: #e8f4fd; }}
        .pct {{ color: #6c757d; font-size: 0.9em; }}
        @media print {{ body {{ margin: 0; }} .no-print {{ display: none; }} }}
    </style>
</head>
<body>
    <h1>LKPM Ready Pack</h1>
    <p class="copy-hint no-print">Salin nilai dari tabel di bawah ke formulir OSS. Semua nilai dalam Rupiah (IDR).</p>

    <div class="header-info">
        <p><strong>Perusahaan:</strong> {company}</p>
        <p><strong>NPWP:</strong> {npwp}</p>
        <p><strong>NIB:</strong> {nib}</p>
        <p><strong>Periode:</strong> {pack.quarter} {pack.year}</p>
        <p><strong>KBLI:</strong> {", ".join(pack.kbli_codes) or "—"}</p>
        <p><strong>Realisasi terhadap Rencana:</strong> {pack.realization_percentage}%</p>
    </div>

    {f'<div class="validation"><strong>Validasi:</strong><br>{validation_html}</div>' if validation_html else ""}

    <h2>1. Realisasi Investasi — Periode Ini ({pack.quarter} {pack.year})</h2>
    <table>
        <thead>
            <tr>
                <th>Kategori</th>
                <th>Dalam Negeri (IDR)</th>
                <th>Impor (IDR)</th>
                <th>Total (IDR)</th>
            </tr>
        </thead>
        <tbody>
            <tr>
                <td>Peralatan/Mesin</td>
                <td class="amount">{format_idr(pack.realized.equipment_domestic)}</td>
                <td class="amount">{format_idr(pack.realized.equipment_import)}</td>
                <td class="amount">{format_idr(pack.realized.equipment_domestic + pack.realized.equipment_import)}</td>
            </tr>
            <tr>
                <td>Bangunan/Gedung</td>
                <td class="amount">{format_idr(pack.realized.building_domestic)}</td>
                <td class="amount">{format_idr(pack.realized.building_import)}</td>
                <td class="amount">{format_idr(pack.realized.building_domestic + pack.realized.building_import)}</td>
            </tr>
            <tr>
                <td>Kendaraan</td>
                <td class="amount">{format_idr(pack.realized.vehicle_domestic)}</td>
                <td class="amount">{format_idr(pack.realized.vehicle_import)}</td>
                <td class="amount">{format_idr(pack.realized.vehicle_domestic + pack.realized.vehicle_import)}</td>
            </tr>
            <tr>
                <td>Tanah</td>
                <td class="amount">{format_idr(pack.realized.land)}</td>
                <td class="amount">—</td>
                <td class="amount">{format_idr(pack.realized.land)}</td>
            </tr>
            <tr>
                <td>Modal Kerja</td>
                <td class="amount">{format_idr(pack.realized.working_capital)}</td>
                <td class="amount">—</td>
                <td class="amount">{format_idr(pack.realized.working_capital)}</td>
            </tr>
            <tr>
                <td>Lain-lain</td>
                <td class="amount">{format_idr(pack.realized.other)}</td>
                <td class="amount">—</td>
                <td class="amount">{format_idr(pack.realized.other)}</td>
            </tr>
            <tr class="total-row">
                <td>TOTAL</td>
                <td class="amount">{format_idr(pack.realized.total_domestic)}</td>
                <td class="amount">{format_idr(pack.realized.total_import)}</td>
                <td class="amount">{format_idr(pack.realized_total)}</td>
            </tr>
        </tbody>
    </table>

    <h2>2. Realisasi Kumulatif</h2>
    <table>
        <thead>
            <tr>
                <th>Kategori</th>
                <th>Kumulatif (IDR)</th>
                <th>Rencana (IDR)</th>
                <th>% Realisasi</th>
            </tr>
        </thead>
        <tbody>
            {_cumulative_row("Peralatan/Mesin DN", pack.cumulative.equipment_domestic, pack.plan.equipment_domestic)}
            {_cumulative_row("Peralatan/Mesin Impor", pack.cumulative.equipment_import, pack.plan.equipment_import)}
            {_cumulative_row("Bangunan DN", pack.cumulative.building_domestic, pack.plan.building_domestic)}
            {_cumulative_row("Bangunan Impor", pack.cumulative.building_import, pack.plan.building_import)}
            {_cumulative_row("Kendaraan DN", pack.cumulative.vehicle_domestic, pack.plan.vehicle_domestic)}
            {_cumulative_row("Kendaraan Impor", pack.cumulative.vehicle_import, pack.plan.vehicle_import)}
            {_cumulative_row("Tanah", pack.cumulative.land, pack.plan.land)}
            {_cumulative_row("Modal Kerja", pack.cumulative.working_capital, pack.plan.working_capital)}
            {_cumulative_row("Lain-lain", pack.cumulative.other, pack.plan.other)}
            <tr class="total-row">
                <td>TOTAL</td>
                <td class="amount">{format_idr(pack.cumulative_total)}</td>
                <td class="amount">{format_idr(pack.plan.grand_total)}</td>
                <td class="amount">{pack.realization_percentage}%</td>
            </tr>
        </tbody>
    </table>

    <h2>3. Tenaga Kerja</h2>
    <table>
        <thead>
            <tr><th>Kategori</th><th>Jumlah</th></tr>
        </thead>
        <tbody>
            <tr><td>TKI (Tenaga Kerja Indonesia)</td><td>{pack.employment.tki}</td></tr>
            <tr><td>TKA (Tenaga Kerja Asing)</td><td>{pack.employment.tka}</td></tr>
            <tr class="total-row"><td>TOTAL</td><td>{pack.employment.total}</td></tr>
        </tbody>
    </table>

    <h2>4. Pendapatan</h2>
    <table>
        <tbody>
            <tr><td>Pendapatan Triwulan</td><td class="amount">Rp {format_idr(pack.quarterly_revenue)}</td></tr>
            <tr><td>Pendapatan Tahunan</td><td class="amount">Rp {format_idr(pack.annual_revenue)}</td></tr>
        </tbody>
    </table>

    <h2>5. Hambatan & Rencana</h2>
    <div style="background:#f8f9fa;padding:12px;border-radius:6px;">
        <p><strong>Hambatan:</strong></p>
        <p>{obstacles}</p>
        <p><strong>Rencana Tindak Lanjut:</strong></p>
        <p>{plans}</p>
    </div>

    <div class="no-print" style="margin-top:24px;padding:12px;background:#e8f4fd;border-radius:6px;">
        <p><strong>Generated:</strong> {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}</p>
        <p><strong>Draft ID:</strong> {pack.draft_id}</p>
    </div>
</body>
</html>"""



def _cumulative_row(label: str, cumulative: int, planned: int) -> str:
    """Generate a cumulative comparison table row."""
    pct = f"{(cumulative / planned * 100):.1f}%" if planned > 0 else "—"
    return (
        f"<tr>"
        f"<td>{label}</td>"
        f'<td class="amount">{format_idr(cumulative)}</td>'
        f'<td class="amount">{format_idr(planned)}</td>'
        f'<td class="amount">{pct}</td>'
        f"</tr>"
    )


# ── LkpmReadyPack ─────────────────────────────────────────────────────────


class LkpmReadyPack:
    """
    Orchestrate LKPM ready-pack generation:

    1. Validate completeness via LKPMValidator.check_completeness_async.
    2. Fetch client + lkpm_reports + lkpm_receipts.
    3. Build PDF (reportlab) + XLSX (openpyxl) in-memory.
    4. If not dry_run: upload to Drive (graceful — failure → drive_url=None).
    5. If send_email + drive_url + client email: send via Brevo (graceful).

    Returns:
        dict with keys: drive_url, pdf_sha256, xlsx_sha256,
                        email_sent_to, validation_warnings.
    """

    def __init__(
        self,
        db_pool: "asyncpg.Pool | None" = None,
        *,
        drive: "Any | None" = None,
        brevo: "Any | None" = None,
        connection: "asyncpg.Connection | None" = None,
    ) -> None:
        self._pool = db_pool
        self._conn = connection
        self._drive = drive
        self._brevo = brevo

    @classmethod
    def with_connection(
        cls,
        conn: "asyncpg.Connection",
        *,
        drive: "Any | None" = None,
        brevo: "Any | None" = None,
    ) -> "LkpmReadyPack":
        """Factory for when a caller already holds a connection (e.g. inside a transaction)."""
        instance = cls(db_pool=None, drive=drive, brevo=brevo, connection=conn)
        return instance

    async def _acquire(self) -> "asyncpg.Connection":
        """Return the shared connection or acquire one from the pool."""
        if self._conn is not None:
            return self._conn
        if self._pool is not None:
            return await self._pool.acquire()  # type: ignore[return-value]
        raise RuntimeError("LkpmReadyPack: neither db_pool nor connection provided")

    async def generate(
        self,
        client_id: int,
        period: str,
        send_email: bool = True,
        dry_run: bool = False,
    ) -> dict:
        """
        Generate a LKPM ready-pack.

        Args:
            client_id: CRM clients.id
            period: "Q1 2026" format
            send_email: whether to send Brevo email after upload
            dry_run: if True, skip Drive upload and email, return hashes only

        Returns:
            {
                "drive_url": str | None,
                "pdf_sha256": str,
                "xlsx_sha256": str,
                "email_sent_to": str | None,
                "validation_warnings": list[str],
            }

        Raises:
            LkpmValidationError: if the report is incomplete.
        """
        import hashlib

        import asyncpg

        from backend.services.compliance.exceptions import LkpmValidationError
        from backend.services.compliance.lkpm_validator import LKPMValidator

        # ── Step 1: Completeness check + Step 2: Fetch data ──────────
        # Wrap in try/finally so pool-acquired connections are always released,
        # even when LkpmValidationError is raised mid-flight (Critical fix: pool leak).
        conn = await self._acquire()
        try:
            validator = LKPMValidator(self._pool)  # type: ignore[arg-type]
            completeness = await validator.check_completeness_async(
                conn,
                client_id,
                period,
            )
            if not completeness["is_complete"]:
                raise LkpmValidationError(
                    f"LKPM report for client_id={client_id}, period={period!r} "
                    f"is incomplete. Missing: {completeness['missing_fields']}"
                )
            validation_warnings: list[str] = completeness.get("warnings", [])

            # ── Step 2: Fetch data ─────────────────────────────────────────
            parts = period.strip().split()
            quarter, year = parts[0], int(parts[1])

            # Fetch client row
            client_row = await conn.fetchrow(
                """
                SELECT id, full_name, email, google_drive_folder_id
                FROM clients
                WHERE id = $1
                """,
                client_id,
            )
            if client_row is None:
                raise LkpmValidationError(f"Client {client_id} not found")

            client_name: str = client_row["full_name"] or f"client_{client_id}"
            client_email: str | None = client_row.get("email") or None
            client_drive_folder: str | None = client_row.get("google_drive_folder_id") or None

            # Fetch lkpm_reports row + NIB via lkpm_client_config
            report_row = await conn.fetchrow(
                """
                SELECT r.id AS report_id, r.status, r.lkpm_assigned_to,
                       r.realized_equipment_domestic + r.realized_equipment_import +
                       r.realized_building_domestic + r.realized_building_import +
                       r.realized_vehicle_domestic + r.realized_vehicle_import +
                       r.realized_land + r.realized_working_capital + r.realized_other
                       AS realization_total,
                       COALESCE(cfg.nib, '') AS nib
                FROM lkpm_reports r
                LEFT JOIN lkpm_client_config cfg
                    ON cfg.client_id = r.client_id
                   AND COALESCE(r.company_id, 0) = COALESCE(cfg.company_id, 0)
                WHERE r.client_id = $1 AND r.quarter = $2 AND r.year = $3
                LIMIT 1
                """,
                client_id,
                quarter,
                year,
            )
            if report_row is None:
                raise LkpmValidationError(
                    f"No lkpm_reports row for client_id={client_id}, quarter={quarter}, year={year}"
                )

            report_id: int = report_row["report_id"]
            assignee: str = report_row.get("lkpm_assigned_to") or "—"
            nib: str = report_row.get("nib") or "—"
            realization_idr: int = int(report_row.get("realization_total") or 0)

            # Fetch lkpm_receipts
            receipt_rows = await conn.fetch(
                """
                SELECT kbli_code, kegiatan_usaha_desc, oss_status, stage
                FROM lkpm_receipts
                WHERE lkpm_report_id = $1
                ORDER BY id
                """,
                report_id,
            )
            kbli_rows: list[dict] = [dict(r) for r in receipt_rows]

        finally:
            # Release only when we acquired from the pool (not a pre-existing connection)
            if self._conn is None and self._pool is not None:
                await self._pool.release(conn)

        # ── Step 3: Build PDF + XLSX ───────────────────────────────────
        from backend.services.compliance.lkpm_pdf_builder import (
            LkpmPackData,
            LkpmPdfBuilder,
        )
        from datetime import datetime, timezone

        pack_data = LkpmPackData(
            client_name=client_name,
            pt_nib=nib,
            period=period,
            kbli_rows=kbli_rows,
            assignee=assignee,
            generated_at=datetime.now(timezone.utc),
            realization_idr=max(0, realization_idr),
        )
        pdf_bytes = LkpmPdfBuilder().build(pack_data)
        xlsx_bytes = self._build_xlsx(pack_data)

        pdf_sha256 = hashlib.sha256(pdf_bytes).hexdigest()
        xlsx_sha256 = hashlib.sha256(xlsx_bytes).hexdigest()

        if dry_run:
            return {
                "drive_url": None,
                "pdf_sha256": pdf_sha256,
                "xlsx_sha256": xlsx_sha256,
                "email_sent_to": None,
                "validation_warnings": validation_warnings,
            }

        # ── Step 4: Upload to Drive ────────────────────────────────────
        drive_url: str | None = None
        if self._drive is not None and client_drive_folder:
            try:
                file_name = f"LKPM_{period.replace(' ', '_')}_{client_id}.pdf"
                result = await self._drive.upload_file_to_folder(
                    user_id="system",
                    folder_id=client_drive_folder,
                    file_content=pdf_bytes,
                    file_name=file_name,
                    mime_type="application/pdf",
                )
                drive_url = result.get("download_url") or result.get("id")
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "LkpmReadyPack: Drive upload failed for client %d / %s — %s",
                    client_id,
                    period,
                    exc,
                )
                drive_url = None

        # ── Step 5: Send email via Brevo ───────────────────────────────
        email_sent_to: str | None = None
        if send_email and drive_url and client_email:
            try:
                from backend.services.compliance.templates_i18n import render_template

                # Default to English — preferred_language column may not exist
                lang = "en"
                subject = render_template("lkpm", "readypack_subject", lang, period=period)
                body = render_template(
                    "lkpm", "readypack_body", lang, period=period, drive_url=drive_url
                )

                if self._brevo is not None:
                    # Use injected service (e.g. in tests)
                    success = await self._brevo.send(
                        to=client_email,
                        subject=subject,
                        body=body,
                        from_email="zantara@balizero.com",
                        from_name="Zantara",
                    )
                else:
                    # Direct Brevo HTTP call (same pattern as notifications/router.py)
                    success = await self._send_brevo_direct(
                        to_email=client_email,
                        subject=subject,
                        body=body,
                    )

                if success:
                    email_sent_to = client_email
                else:
                    logger.warning(
                        "LkpmReadyPack: Brevo send returned False for %s", client_email
                    )
            except Exception as exc:  # noqa: BLE001
                logger.warning(
                    "LkpmReadyPack: email send failed for client %d — %s", client_id, exc
                )
                email_sent_to = None

        return {
            "drive_url": drive_url,
            "pdf_sha256": pdf_sha256,
            "xlsx_sha256": xlsx_sha256,
            "email_sent_to": email_sent_to,
            "validation_warnings": validation_warnings,
        }

    # ── Private helpers ────────────────────────────────────────────────────

    @staticmethod
    def _build_xlsx(data: "Any") -> bytes:
        """Build an XLSX file in memory from LkpmPackData."""
        import io as _io

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"LKPM {data.period}"

        header_fill = PatternFill("solid", fgColor="1a5276")
        header_font = Font(bold=True, color="FFFFFF")

        # Metadata block
        ws.append(["LKPM Ready Pack"])
        ws.append(["Perusahaan", data.client_name])
        ws.append(["NIB", data.pt_nib])
        ws.append(["Periode", data.period])
        ws.append(["Assignee", data.assignee])
        ws.append(["Realisasi (IDR)", data.realization_idr])
        ws.append(["Dibuat", data.generated_at.strftime("%Y-%m-%d %H:%M UTC")])
        ws.append([])  # blank row

        # KBLI header
        kbli_header = ["Kode KBLI", "Kegiatan Usaha", "Status OSS"]
        ws.append(kbli_header)
        header_row_idx = ws.max_row
        for col_idx, _ in enumerate(kbli_header, start=1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        # KBLI rows
        if data.kbli_rows:
            for row in data.kbli_rows:
                ws.append(
                    [
                        row.get("kbli_code", ""),
                        row.get("kegiatan_usaha_desc", ""),
                        row.get("oss_status", ""),
                    ]
                )
        else:
            ws.append(["—", "Tidak ada baris KBLI.", "—"])

        buf = _io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @staticmethod
    async def _send_brevo_direct(to_email: str, subject: str, body: str) -> bool:
        """Send a transactional email via Brevo HTTP API."""
        import os

        api_key = os.getenv("SENDGRID_API_KEY", "")
        if not api_key:
            logger.warning("LkpmReadyPack: SENDGRID_API_KEY not set — skip email")
            return False

        is_brevo = api_key.startswith("xkeysib-")
        url = (
            "https://api.brevo.com/v3/smtp/email"
            if is_brevo
            else "https://api.sendgrid.com/v3/mail/send"
        )
        headers = (
            {"api-key": api_key, "Content-Type": "application/json"}
            if is_brevo
            else {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
        )
        payload = {
            "sender": {"email": "zantara@balizero.com", "name": "Zantara"},
            "to": [{"email": to_email}],
            "subject": subject,
            "htmlContent": body,
        }

        try:
            client = _get_brevo_client()
            resp = await client.post(url, json=payload, headers=headers)
            if resp.status_code in (200, 201, 202):
                return True
            logger.error(
                "LkpmReadyPack: Brevo API error %d: %s",
                resp.status_code,
                resp.text[:300],
            )
            return False
        except httpx.HTTPError as exc:
            logger.error("LkpmReadyPack: HTTP error sending email — %s", exc)
            return False


__all__ = [
    # original exports
    "OBSTACLE_TEMPLATES",
    "format_idr",
    "generate_ready_pack_html",
    # new exports
    "LkpmReadyPack",
    "close_brevo_client",
]
